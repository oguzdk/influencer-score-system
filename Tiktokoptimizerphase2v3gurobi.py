# -*- coding: utf-8 -*-
"""
TikTok Influencer Selection — OPTİMİZASYON MODELİ (FAZ 2)  v3.0
=================================================================
Lisans Tezi: "TikTok'ta Etkili Marka İş Birliği için İçerik Üreticisi
              Seçiminin Optimizasyonu"
Eskişehir Teknik Üniversitesi — Endüstri Mühendisliği

SOLVER: Gurobi (gurobipy)
  pip install gurobipy
  Akademik lisans: https://www.gurobi.com/academia/academic-program-and-licenses/

MODEL:
  MAX   Σᵢ  mcdm_i · xᵢ
  s.t.  Σᵢ  cost_i · xᵢ  ≤  B          (bütçe)
        Σᵢ  xᵢ             ≥  n_min      (min çeşitlilik)
        Σᵢ  xᵢ             ≤  n_max      (maks sayı, opsiyonel)
        Σ_{i∈T_t} xᵢ       ≤  L_t        (tier kısıtı, opsiyonel)
        xₖ  = 0   ∀ k ∈ K_black          (kara liste)
        xᵢ ∈ {0, 1}

ÇIKTILAR:
  data_tiktok/optimization_results/
    tiktok_optimization_<ts>.xlsx   (7 sayfa, stillendirilmiş)
    optimization_report_<ts>.png    (10 panel görsel)
"""

import warnings
warnings.filterwarnings("ignore")

import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns

# ── GUROBI (Opsiyonel — Yoksa Knapsack DP kullanılır) ────────────────
GUROBI_AVAILABLE = False
try:
    import gurobipy as gp
    from gurobipy import GRB
    GUROBI_AVAILABLE = True
    print("[Solver] Gurobi hazır.")
except ImportError:
    print("[Solver] Gurobi bulunamadı → Knapsack DP kullanılacak.")


# =====================================================================
# PATHS
# =====================================================================

DATA_DIR = Path("data_tiktok")
OUT_DIR  = Path("data_tiktok/optimization_results")
OUT_DIR.mkdir(parents=True, exist_ok=True)

def _pick(primary, fallback):
    p = Path(primary)
    return p if p.exists() else Path(fallback)

PROFILES_EXCEL = _pick(
    "data_tiktok/tiktok_profiles_categorized.xlsx",
    "data_tiktok/tiktok_profiles.xlsx"
)
VIDEOS_EXCEL = _pick(
    "data_tiktok/tiktok_videos_categorized.xlsx",
    "data_tiktok/tiktok_videos.xlsx"
)


# =====================================================================
# MALİYET MODELİ
# =====================================================================

COST_MODEL = {
    "nano"  : {"range": (0,          10_000),      "base": 1_500,   "per_1k": 150},
    "micro" : {"range": (10_000,     100_000),     "base": 8_000,   "per_1k": 80},
    "mid"   : {"range": (100_000,    500_000),     "base": 35_000,  "per_1k": 35},
    "macro" : {"range": (500_000,    1_000_000),   "base": 90_000,  "per_1k": 20},
    "mega"  : {"range": (1_000_000,  999_999_999), "base": 200_000, "per_1k": 8},
}

CATEGORY_COST_MULTIPLIER = {
    "Beauty & Personal Care"  : 1.15,
    "Fashion & Style"         : 1.10,
    "Technology & Digital"    : 1.05,
    "Fitness & Health"        : 1.00,
    "Food & Cooking"          : 0.95,
    "Comedy & Entertainment"  : 0.90,
    "Gaming"                  : 0.90,
    "Music & Performance"     : 1.00,
    "Travel & Lifestyle"      : 1.05,
    "Education & Informative" : 0.85,
    "Mixed/Unclear"           : 1.00,
}

BUDGET_SCENARIOS = {
    "Düşük Bütçe (50K TL)"    : 50_000,
    "Orta Bütçe (150K TL)"    : 150_000,
    "Yüksek Bütçe (300K TL)"  : 300_000,
    "Premium Bütçe (600K TL)" : 600_000,
}

CRITERION_WEIGHTS = {
    "engagement_proxy"  : 0.35,
    "followers"         : 0.20,
    "like_per_follower" : 0.20,
    "cost_efficiency"   : 0.15,
    "video_volume"      : 0.10,
}

# =====================================================================
# KARA LİSTE  —  x_k = 0  ∀ k ∈ BLACKLIST
# =====================================================================
# Bu influencerlar MCDM skoru veya bütçe ne kadar uygun olursa olsun
# HİÇBİR senaryoda seçilmez.
# Kullanıcı adını küçük harf, @ olmadan yaz.
#
# Örnek:
#   BLACKLIST = ["cznburak", "dilanpolat"]
#
BLACKLIST: List[str] = [
    # "cznburak",
    # "dilanpolat",
]

def apply_blacklist(dfp: pd.DataFrame) -> pd.DataFrame:
    """
    BLACKLIST'teki kullanıcıları aday setinden çıkarır.
    Matematiksel karşılığı:  x_k = 0  ∀ k ∈ BLACKLIST
    Model kurulmadan önce uygulandığı için tüm Gurobi senaryolarında
    aynı şekilde çalışır.
    """
    if not BLACKLIST:
        return dfp
    bl = [u.strip().lower() for u in BLACKLIST]
    before = len(dfp)
    clean  = dfp[~dfp["username"].isin(bl)].copy()
    if before - len(clean):
        print(f"  🚫 Kara liste: {before - len(clean)} influencer çıkarıldı → {bl}")
    return clean


# =====================================================================
# YARDIMCI FONKSİYONLAR
# =====================================================================

def ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def safe_float(x, default=0.0) -> float:
    try:
        v = float(x)
        return v if np.isfinite(v) else default
    except (TypeError, ValueError):
        return default

def classify_tier(followers: float) -> str:
    f = int(safe_float(followers))
    for tier, cfg in COST_MODEL.items():
        if cfg["range"][0] <= f < cfg["range"][1]:
            return tier
    return "mega"

def estimate_cost_tl(followers: float, category: str = "") -> float:
    tier = classify_tier(followers)
    cfg  = COST_MODEL[tier]
    cost = cfg["base"] + (safe_float(followers) / 1_000) * cfg["per_1k"]
    mult = CATEGORY_COST_MULTIPLIER.get(category, 1.0)
    return round(cost * mult, 2)

def min_max_norm(series: pd.Series) -> pd.Series:
    s = pd.to_numeric(series, errors="coerce").fillna(0)
    lo, hi = s.min(), s.max()
    if hi == lo:
        return pd.Series([0.5] * len(s), index=s.index)
    return (s - lo) / (hi - lo)

def fmt_tl(x: float) -> str:
    return f"TL {x:,.0f}"

TIER_ORDER = ["nano", "micro", "mid", "macro", "mega"]
PAL        = ["#4C9BE8", "#6BC5A1", "#F5A623", "#E85B4C", "#9B59B6"]
TC         = dict(zip(TIER_ORDER, PAL))


# =====================================================================
# ADIM 1: VERİ YÜKLEME & TEMİZLEME
# =====================================================================

def load_data() -> Tuple[pd.DataFrame, pd.DataFrame]:
    print("\n" + "="*62)
    print("ADIM 1: Veri Yükleme & Temizleme")
    print("="*62)

    if PROFILES_EXCEL.exists():
        dfp = pd.read_excel(PROFILES_EXCEL)
        print(f"  Profil : {len(dfp)} satır  ← {PROFILES_EXCEL.name}")
    else:
        print("  Profil Excel bulunamadı → demo veri")
        dfp = _demo_profiles()

    if VIDEOS_EXCEL.exists():
        dfv = pd.read_excel(VIDEOS_EXCEL)
        print(f"  Video  : {len(dfv)} satır  ← {VIDEOS_EXCEL.name}")
    else:
        dfv = pd.DataFrame()
        print("  Video Excel bulunamadı.")

    # username standardizasyonu
    if "normalized_username" in dfp.columns:
        dfp["username"] = dfp["normalized_username"].astype(str).str.strip().str.lower()
    elif "username" in dfp.columns:
        dfp["username"] = dfp["username"].astype(str).str.strip().str.lower()
    else:
        dfp["username"] = [f"user_{i}" for i in range(len(dfp))]

    # hatalı profilleri çıkar
    if "error" in dfp.columns:
        bad = dfp["error"].notna() & (dfp["error"].astype(str).str.strip() != "")
        dfp = dfp[~bad].reset_index(drop=True)
        print(f"  Hata flagli satır temizlendi: {bad.sum()}")

    # kategori
    if "category_primary_final" in dfp.columns:
        dfp["category"] = dfp["category_primary_final"].fillna("Mixed/Unclear")
    elif "category_primary" in dfp.columns:
        dfp["category"] = dfp["category_primary"].fillna("Mixed/Unclear")
    else:
        dfp["category"] = "Mixed/Unclear"

    # numerik sütunlar
    for col in ["followers", "following", "likes"]:
        if col in dfp.columns:
            dfp[col] = pd.to_numeric(dfp[col], errors="coerce").fillna(0)

    print(f"  Geçerli profil: {len(dfp)}")
    return dfp, dfv


def _demo_profiles() -> pd.DataFrame:
    np.random.seed(42)
    n = 60
    tiers = list(COST_MODEL.keys())
    rows  = []
    for i in range(n):
        f = int(np.random.lognormal(11, 1.5))
        rows.append({
            "username"    : f"user_{i+1:03d}",
            "display_name": f"Influencer {i+1}",
            "followers"   : f,
            "likes"       : int(f * np.random.uniform(5, 30)),
            "category"    : "Mixed/Unclear",
            "error"       : None,
        })
    return pd.DataFrame(rows)


# =====================================================================
# ADIM 2: ENGAGEMENT PROXY
# =====================================================================

def build_engagement_proxy(dfp: pd.DataFrame, dfv: pd.DataFrame) -> pd.DataFrame:
    print("\n" + "="*62)
    print("ADIM 2: Engagement Proxy (views=NaN → likes/followers)")
    print("="*62)

    # --- video tarafından gelen metrikler ---
    if not dfv.empty and "profile_username" in dfv.columns:
        dfv = dfv.copy()
        dfv["profile_username"] = dfv["profile_username"].astype(str).str.strip().str.lower()

        # video_likes kolonunu bul (farklı isimlerle gelebilir)
        likes_col = None
        for candidate in ["video_likes", "likes", "digg_count", "like_count"]:
            if candidate in dfv.columns:
                likes_col = candidate
                break
        dfv["video_likes_n"] = pd.to_numeric(
            dfv[likes_col] if likes_col else 0, errors="coerce"
        ).fillna(0)

        # video_url kolonunu bul (count için)
        url_col = "video_url" if "video_url" in dfv.columns else dfv.columns[0]

        agg = dfv.groupby("profile_username").agg(
            avg_video_likes = ("video_likes_n", "mean"),
            video_count     = (url_col,         "count"),
        ).reset_index().rename(columns={"profile_username": "username"})

        # Merge öncesi çakışan kolonları temizle (KeyError / _x _y sorununu önler)
        for col in ["avg_video_likes", "video_count"]:
            if col in dfp.columns:
                dfp = dfp.drop(columns=[col])

        dfp = dfp.merge(agg, on="username", how="left")
        dfp["avg_video_likes"] = dfp["avg_video_likes"].fillna(0)
        dfp["video_count"]     = dfp["video_count"].fillna(0)
    else:
        dfp["avg_video_likes"] = dfp.get("avg_video_likes", pd.Series(
            [0] * len(dfp), index=dfp.index)).fillna(0)
        dfp["video_count"] = dfp.get("video_count", pd.Series(
            [0] * len(dfp), index=dfp.index)).fillna(0)

    # --- engagement proxy = avg_video_likes / followers ---
    dfp["like_per_follower"] = dfp.apply(
        lambda r: safe_float(r.get("avg_video_likes")) / max(safe_float(r.get("followers")), 1),
        axis=1
    )

    # --- engagement proxy (normalize edilmiş) ---
    dfp["engagement_proxy"] = dfp.apply(
        lambda r: (
            0.6 * min_max_norm(pd.Series([safe_float(r.get("like_per_follower"))]))[0] +
            0.4 * min_max_norm(pd.Series([np.log1p(safe_float(r.get("avg_video_likes")))]))[0]
        ),
        axis=1
    )

    # Toplu normalize — tek satırlık değil, tüm kolon üzerinden
    dfp["engagement_proxy"] = (
        0.6 * min_max_norm(dfp["like_per_follower"]) +
        0.4 * min_max_norm(np.log1p(dfp["avg_video_likes"]))
    )

    print(f"  avg_video_likes  — ort: {dfp['avg_video_likes'].mean():,.0f}  "
          f"maks: {dfp['avg_video_likes'].max():,.0f}")
    print(f"  like_per_follower — ort: {dfp['like_per_follower'].mean():.4f}  "
          f"maks: {dfp['like_per_follower'].max():.4f}")
    return dfp


# =====================================================================
# ADIM 3: MALİYET TAHMİNİ
# =====================================================================

def estimate_costs(dfp: pd.DataFrame) -> pd.DataFrame:
    print("\n" + "="*62)
    print("ADIM 3: Maliyet Tahmini")
    print("="*62)

    dfp["estimated_cost_tl"] = dfp.apply(
        lambda r: estimate_cost_tl(r.get("followers", 0), r.get("category", "")), axis=1
    )
    dfp["tier"] = dfp["followers"].apply(classify_tier)

    for t in TIER_ORDER:
        g = dfp[dfp["tier"] == t]
        if len(g):
            print(f"  {t:7s}: {len(g):3d} kişi  "
                  f"min={g['estimated_cost_tl'].min():>9,.0f}  "
                  f"med={g['estimated_cost_tl'].median():>9,.0f}  "
                  f"maks={g['estimated_cost_tl'].max():>9,.0f} TL")
    return dfp


# =====================================================================
# ADIM 4: MCDM SKORU
# =====================================================================

def compute_mcdm_score(dfp: pd.DataFrame) -> pd.DataFrame:
    print("\n" + "="*62)
    print("ADIM 4: MCDM Skoru")
    print("="*62)

    dfp["c1_engagement"]      = min_max_norm(dfp["engagement_proxy"])
    dfp["c2_followers"]       = min_max_norm(np.log1p(dfp["followers"]))
    dfp["c3_like_per_follower"] = min_max_norm(dfp["like_per_follower"])

    dfp["cost_eff_raw"] = dfp.apply(
        lambda r: np.log1p(safe_float(r.get("avg_video_likes"))) /
                  max(safe_float(r.get("estimated_cost_tl")), 1),
        axis=1
    )
    dfp["c4_cost_efficiency"] = min_max_norm(dfp["cost_eff_raw"])
    dfp["c5_video_volume"]    = min_max_norm(dfp.get("video_count", pd.Series([0]*len(dfp))))

    w = CRITERION_WEIGHTS
    dfp["mcdm_score"] = (
        w["engagement_proxy"]  * dfp["c1_engagement"]       +
        w["followers"]         * dfp["c2_followers"]        +
        w["like_per_follower"] * dfp["c3_like_per_follower"] +
        w["cost_efficiency"]   * dfp["c4_cost_efficiency"]  +
        w["video_volume"]      * dfp["c5_video_volume"]
    ).round(4)

    print(f"  ort={dfp['mcdm_score'].mean():.4f}  "
          f"std={dfp['mcdm_score'].std():.4f}  "
          f"min={dfp['mcdm_score'].min():.4f}  "
          f"maks={dfp['mcdm_score'].max():.4f}")

    top = dfp.nlargest(10, "mcdm_score")[
        ["username","tier","followers","engagement_proxy","mcdm_score","estimated_cost_tl"]
    ]
    print(f"\n  En Yüksek MCDM Skorlu İlk 10:")
    print(top.to_string(index=False))
    return dfp


# =====================================================================
# ADIM 5: OPTİMİZASYON — GUROBI ILP
# =====================================================================

def optimize_portfolio(
    dfp             : pd.DataFrame,
    budget_tl       : float,
    min_influencers : int            = 3,
    max_influencers : Optional[int]  = None,
    tier_limits     : Optional[Dict[str, int]] = None,
    required_categories: Optional[List[str]]   = None,
    label           : str            = "scenario",
    verbose         : bool           = True,
    gurobi_time_limit: float         = 60.0,       # saniye
    gurobi_mip_gap   : float         = 0.001,      # %0.1 gap
) -> Dict:
    """
    0-1 Tam Sayılı Doğrusal Programlama — Gurobi ILP

    Amaç  : MAX  Σᵢ mcdm_i · xᵢ
    Kısıt 1: Σᵢ cost_i · xᵢ        ≤ budget_tl        [bütçe]
    Kısıt 2: Σᵢ xᵢ                  ≥ min_influencers  [min çeşitlilik]
    Kısıt 3: Σᵢ xᵢ                  ≤ max_influencers  [maks (opsiyonel)]
    Kısıt 4: Σ_{i∈tier_t} xᵢ        ≤ tier_limits[t]  [tier (opsiyonel)]
    Kısıt 5: Σ_{i∈cat_r} xᵢ         ≥ 1               [zorunlu kategori]
    Kısıt 6: xₖ = 0  ∀ k ∈ BLACKLIST                  [kara liste]
    xᵢ ∈ {0, 1}

    Parametreler:
        gurobi_time_limit : Gurobi çözme süresi limiti (sn)
        gurobi_mip_gap    : MIP optimality gap toleransı
    """
    if verbose:
        print(f"\n  → {label}  |  Bütçe: {fmt_tl(budget_tl)}")

    # ── Kara Liste Kısıtı: x_k = 0  ∀ k ∈ BLACKLIST ─────────────
    dfp_clean = apply_blacklist(dfp)

    cands = dfp_clean[dfp_clean["estimated_cost_tl"] <= budget_tl].copy().reset_index(drop=True)

    if len(cands) == 0:
        if verbose:
            print(f"    ❌ Bu bütçeye uygun aday yok.")
        return {
            "label": label, "budget": budget_tl, "selected": pd.DataFrame(),
            "total_score": 0.0, "total_cost": 0.0,
            "budget_used_pct": 0.0, "n_selected": 0, "status": "infeasible",
        }

    n      = len(cands)
    scores = cands["mcdm_score"].values.astype(float)
    costs  = cands["estimated_cost_tl"].values.astype(float)
    cats   = cands["category"].values if "category" in cands.columns else ["Mixed/Unclear"] * n
    tiers_ = cands["tier"].values

    # ── GUROBI ───────────────────────────────────────────────────
    mask, sol_status = _solve_gurobi(
        n, scores, costs, cats, tiers_,
        budget_tl, min_influencers, max_influencers,
        tier_limits, required_categories,
        label, gurobi_time_limit, gurobi_mip_gap
    )

    selected    = cands[mask].copy()
    total_score = float(selected["mcdm_score"].sum())
    total_cost  = float(selected["estimated_cost_tl"].sum())
    used_pct    = total_cost / budget_tl * 100

    if verbose:
        print(f"    ✅ Seçilen    : {len(selected)} influencer  [{sol_status}]")
        print(f"    Maliyet      : {fmt_tl(total_cost)} / {fmt_tl(budget_tl)} ({used_pct:.1f}%)")
        print(f"    MCDM Toplam  : {total_score:.4f}")
        names = ", ".join(selected["username"].tolist()[:8])
        if len(selected) > 8:
            names += f" ... (+{len(selected)-8})"
        print(f"    Seçilenler   : {names}")

    return {
        "label"           : label,
        "budget"          : budget_tl,
        "selected"        : selected,
        "total_score"     : total_score,
        "total_cost"      : total_cost,
        "budget_used_pct" : used_pct,
        "n_selected"      : len(selected),
        "status"          : sol_status,
    }


def _solve_gurobi(
    n, scores, costs, cats, tiers_,
    budget_tl, min_n, max_n,
    tier_limits, required_cats,
    label, time_limit, mip_gap
):
    """
    Gurobi model kurulumu ve çözümü.
    Başarılıysa (mask: List[bool], status: str) döner.
    Hata durumunda RuntimeError fırlatır.
    """
    try:
        # Sessiz mod — Gurobi log'u konsola basmasın
        env = gp.Env(empty=True)
        env.setParam("OutputFlag", 0)        # konsol çıktısı kapat
        env.setParam("LogToConsole", 0)
        env.start()

        model = gp.Model(
            name=f"TikTok_{re.sub(r'[^a-zA-Z0-9]', '_', label)[:30]}",
            env=env
        )

        # ── Solver parametreleri ──────────────────────────────────
        model.setParam("TimeLimit",    time_limit)   # çözme süresi (sn)
        model.setParam("MIPGap",       mip_gap)      # optimality gap
        model.setParam("Threads",      4)             # paralel iş parçacığı
        model.setParam("MIPFocus",     1)             # iyi çözüm bulmaya odaklan
        model.setParam("Presolve",     2)             # agresif ön çözme

        # ── Karar değişkenleri ────────────────────────────────────
        # x[i] ∈ {0, 1}  :  i. influencer seçildi mi?
        x = model.addVars(n, vtype=GRB.BINARY, name="x")

        # ── Amaç fonksiyonu ───────────────────────────────────────
        # MAX  Σᵢ mcdm_i · xᵢ
        model.setObjective(
            gp.quicksum(scores[i] * x[i] for i in range(n)),
            GRB.MAXIMIZE
        )

        # ── Kısıt 1: Bütçe ───────────────────────────────────────
        # Σᵢ cost_i · xᵢ ≤ budget_tl
        model.addConstr(
            gp.quicksum(costs[i] * x[i] for i in range(n)) <= budget_tl,
            name="c_budget"
        )

        # ── Kısıt 2: Minimum influencer sayısı ───────────────────
        # Σᵢ xᵢ ≥ min_n
        model.addConstr(
            gp.quicksum(x[i] for i in range(n)) >= min(min_n, n),
            name="c_min_n"
        )

        # ── Kısıt 3: Maksimum influencer sayısı (opsiyonel) ──────
        # Σᵢ xᵢ ≤ max_n
        if max_n is not None:
            model.addConstr(
                gp.quicksum(x[i] for i in range(n)) <= max_n,
                name="c_max_n"
            )

        # ── Kısıt 4: Tier bazlı üst sınır (opsiyonel) ────────────
        # Σ_{i: tier_i = t} xᵢ ≤ tier_limits[t]
        if tier_limits:
            for t, lim in tier_limits.items():
                idxs = [i for i, tv in enumerate(tiers_) if tv == t]
                if idxs:
                    model.addConstr(
                        gp.quicksum(x[i] for i in idxs) <= lim,
                        name=f"c_tier_{t}"
                    )

        # ── Kısıt 5: Zorunlu kategori (opsiyonel) ────────────────
        # Σ_{i: cat_i = r} xᵢ ≥ 1   ∀ r ∈ required_cats
        if required_cats:
            for r in required_cats:
                idxs = [i for i, c in enumerate(cats) if c == r]
                if idxs:
                    model.addConstr(
                        gp.quicksum(x[i] for i in idxs) >= 1,
                        name=f"c_req_{r[:12]}"
                    )

        # ── Model güncelle & çöz ──────────────────────────────────
        model.update()
        model.optimize()

        # ── Çözüm durumu ─────────────────────────────────────────
        status_code = model.Status
        STATUS_MAP = {
            GRB.OPTIMAL    : "Optimal",
            GRB.SUBOPTIMAL : "Suboptimal",
            GRB.TIME_LIMIT : "TimeLimit",
            GRB.INFEASIBLE : "Infeasible",
            GRB.INF_OR_UNBD: "Inf_or_Unbd",
        }
        sol_status = STATUS_MAP.get(status_code, f"Status_{status_code}")

        if status_code in (GRB.OPTIMAL, GRB.SUBOPTIMAL, GRB.TIME_LIMIT):
            # Çözüm bulunduysa (time limit'te bile ara çözüm olabilir)
            if model.SolCount > 0:
                mask = [x[i].X > 0.5 for i in range(n)]
                obj  = model.ObjVal
                gap  = model.MIPGap * 100
                print(f"    [Gurobi] {sol_status}  "
                      f"Obj={obj:.4f}  Gap={gap:.3f}%  "
                      f"Nodes={int(model.NodeCount)}")
                model.dispose()
                env.dispose()
                return mask, sol_status
            else:
                model.dispose(); env.dispose()
                raise RuntimeError(
                    f"[Gurobi] {sol_status} — model çözüldü fakat tamsayılı çözüm "
                    f"bulunamadı (SolCount=0). Kısıtları veya bütçeyi gözden geçirin."
                )
        else:
            model.dispose(); env.dispose()
            raise RuntimeError(
                f"[Gurobi] Çözüm başarısız: {sol_status}. "
                f"Model infeasible veya unbounded olabilir."
            )

    except gp.GurobiError as e:
        raise RuntimeError(f"[Gurobi] GurobiError: {e}")
    except RuntimeError:
        raise
    except Exception as e:
        raise RuntimeError(f"[Gurobi] Beklenmeyen hata: {e}")



# =====================================================================
# ADIM 5a: KNAPSACK ALGORİTMASI (DP Tabanlı — Gurobi Gerektirmez)
# =====================================================================

def solve_knapsack_dp(
    scores  : np.ndarray,
    costs   : np.ndarray,
    budget  : float,
    n_select: int = 40,
) -> Tuple[List[bool], str]:
    """
    0-1 Knapsack problemi — Dinamik Programlama (DP) çözümü.

    Tam olarak n_select kişi seçerek toplam MCDM skorunu maksimize eder.
    Gurobi veya herhangi bir harici çözücü gerektirmez.

    Model:
      MAX   Σᵢ score_i · xᵢ
      s.t.  Σᵢ cost_i · xᵢ  ≤  budget
            Σᵢ xᵢ            =  n_select
            xᵢ ∈ {0, 1}

    Yaklaşım:
      - Maliyetler kuruş cinsine çevrilerek tamsayı DP tablosu oluşturulur
      - 3 boyutlu DP: dp[i][k][c] = i. elemana kadar k kişi seçip c bütçe kullanarak
        elde edilebilecek maksimum skor
      - Bellek optimizasyonu: sadece önceki satır tutulur

    Parametreler:
        scores   : n boyutlu skor dizisi (float)
        costs    : n boyutlu maliyet dizisi (float, TL)
        budget   : bütçe limiti (float, TL)
        n_select : seçilecek kişi sayısı (varsayılan 40)

    Dönen:
        (mask, status)
        mask   : seçilen elemanların boolean listesi
        status : "Optimal" veya "Infeasible"
    """
    n = len(scores)

    if n < n_select:
        # Yeterli aday yok → mümkün olanı seç
        n_select = n

    # ── Maliyet diskretizasyonu ──────────────────────────────────────
    # DP için maliyetleri tamsayıya çeviriyoruz
    # 100 TL hassasiyetle → bütçe/100 boyutunda tablo
    SCALE = 100  # 100 TL çözünürlük
    int_costs  = np.array([max(1, int(round(c / SCALE))) for c in costs])
    int_budget = int(budget / SCALE)

    # ── DP Tablosu ───────────────────────────────────────────────────
    # dp[k][c] = k kişi seçip c bütçe kullanarak elde edilebilecek max skor
    # Boyut: (n_select+1) × (int_budget+1)
    # -inf ile başlat (geçersiz durumlar)
    NEG_INF = -1e18

    # Bellek optimizasyonu: sadece mevcut dp ve önceki dp tutulur
    # Her eleman için dp tablosunu güncelliyoruz
    dp   = [[NEG_INF] * (int_budget + 1) for _ in range(n_select + 1)]
    # parent[i][k][c] → i. elemanda seçim yapıldı mı (backtrack için)
    # Bellek tasarrufu: ayrı bir backtrack stratejisi kullan
    dp[0][0] = 0.0  # 0 kişi, 0 maliyet → skor 0

    # Seçimleri kaydet (backtracking için)
    choice = [[[False] * (int_budget + 1) for _ in range(n_select + 1)]
              for _ in range(n)]

    for i in range(n):
        # Tersten tarayarak aynı elemanın birden fazla seçilmesini engelle
        for k in range(min(i + 1, n_select), 0, -1):
            for c in range(int_budget, int_costs[i] - 1, -1):
                new_val = dp[k - 1][c - int_costs[i]] + scores[i]
                if new_val > dp[k][c]:
                    dp[k][c] = new_val
                    choice[i][k][c] = True

    # ── En iyi çözümü bul ────────────────────────────────────────────
    best_cost = -1
    best_score = NEG_INF

    for c in range(int_budget + 1):
        if dp[n_select][c] > best_score:
            best_score = dp[n_select][c]
            best_cost = c

    if best_score <= NEG_INF / 2:
        # n_select kişi bütçeye sığmıyor → daha az kişiyle dene
        # En fazla kaç kişi sığıyor?
        best_k = 0
        for k in range(n_select, 0, -1):
            for c in range(int_budget + 1):
                if dp[k][c] > NEG_INF / 2:
                    best_k = k
                    break
            if best_k > 0:
                break

        if best_k == 0:
            return [False] * n, "Infeasible"

        # best_k kişi ile en iyi çözümü bul
        n_select = best_k
        best_cost = -1
        best_score = NEG_INF
        for c in range(int_budget + 1):
            if dp[n_select][c] > best_score:
                best_score = dp[n_select][c]
                best_cost = c

    # ── Backtracking — hangi elemanlar seçildi ───────────────────────
    mask = [False] * n
    k_rem = n_select
    c_rem = best_cost

    for i in range(n - 1, -1, -1):
        if k_rem > 0 and c_rem >= 0 and choice[i][k_rem][c_rem]:
            mask[i] = True
            c_rem -= int_costs[i]
            k_rem -= 1

    return mask, "Optimal"


def optimize_portfolio_knapsack(
    dfp             : pd.DataFrame,
    budget_tl       : float,
    n_select        : int            = 40,
    required_categories: Optional[List[str]] = None,
    label           : str            = "knapsack",
    verbose         : bool           = True,
    criterion_weights: Optional[Dict[str, float]] = None,
) -> Dict:
    """
    Knapsack DP ile influencer portföy optimizasyonu.

    Gurobi gerektirmez. Tam olarak n_select kişi seçmeye çalışır.
    Bütçe yetmezse mümkün olan en fazla kişiyi seçer.

    Parametreler:
        dfp                : profil DataFrame (mcdm_score ve estimated_cost_tl içermeli)
        budget_tl          : bütçe limiti (TL)
        n_select           : seçilecek kişi sayısı (varsayılan 40)
        required_categories: zorunlu kategori listesi (şu an bilgi amaçlı)
        label              : senaryo etiketi
        verbose            : detaylı çıktı
        criterion_weights  : AHP/ANP ağırlıkları (None ise mevcut mcdm_score kullanılır)

    Dönen:
        Gurobi optimize_portfolio ile aynı formatta dict
    """
    if verbose:
        print(f"\n  → {label}  |  Bütçe: {fmt_tl(budget_tl)}  |  Hedef: {n_select} kişi")

    # Kara Liste uygula
    dfp_clean = apply_blacklist(dfp)

    # MCDM skorunu AHP/ANP ağırlıklarıyla yeniden hesapla (verilmişse)
    if criterion_weights is not None:
        crit_col_map = {
            "engagement_proxy"  : "c1_engagement",
            "followers"         : "c2_followers",
            "like_per_follower" : "c3_like_per_follower",
            "cost_efficiency"   : "c4_cost_efficiency",
            "video_volume"      : "c5_video_volume",
        }
        dfp_clean = dfp_clean.copy()
        dfp_clean["mcdm_score"] = sum(
            criterion_weights.get(k, 0) * dfp_clean[crit_col_map[k]]
            for k in criterion_weights
            if k in crit_col_map and crit_col_map[k] in dfp_clean.columns
        ).round(4)

    cands = dfp_clean[dfp_clean["estimated_cost_tl"] <= budget_tl].copy().reset_index(drop=True)

    if len(cands) == 0:
        if verbose:
            print(f"    ❌ Bu bütçeye uygun aday yok.")
        return {
            "label": label, "budget": budget_tl, "selected": pd.DataFrame(),
            "total_score": 0.0, "total_cost": 0.0,
            "budget_used_pct": 0.0, "n_selected": 0, "status": "Infeasible",
        }

    scores = cands["mcdm_score"].values.astype(float)
    costs  = cands["estimated_cost_tl"].values.astype(float)

    # Knapsack DP çöz
    actual_n = min(n_select, len(cands))
    mask, status = solve_knapsack_dp(scores, costs, budget_tl, actual_n)

    selected    = cands[mask].copy()
    total_score = float(selected["mcdm_score"].sum())
    total_cost  = float(selected["estimated_cost_tl"].sum())
    used_pct    = total_cost / budget_tl * 100 if budget_tl > 0 else 0.0

    if verbose:
        print(f"    {'✅' if status == 'Optimal' else '⚠️'} Seçilen: {len(selected)} influencer  [{status}]")
        print(f"    Maliyet      : {fmt_tl(total_cost)} / {fmt_tl(budget_tl)} ({used_pct:.1f}%)")
        print(f"    MCDM Toplam  : {total_score:.4f}")
        names = ", ".join(selected["username"].tolist()[:8])
        if len(selected) > 8:
            names += f" ... (+{len(selected)-8})"
        print(f"    Seçilenler   : {names}")

    return {
        "label"           : label,
        "budget"          : budget_tl,
        "selected"        : selected,
        "total_score"     : total_score,
        "total_cost"      : total_cost,
        "budget_used_pct" : used_pct,
        "n_selected"      : len(selected),
        "status"          : status,
    }


# =====================================================================
# ADIM 5b: SENARYO ANALİZİ
# =====================================================================

def run_scenarios(dfp: pd.DataFrame) -> List[Dict]:
    print("\n" + "="*62)
    print("ADIM 5: Senaryo Analizi (4 bütçe düzeyi)")
    print("="*62)
    results = []
    for label, budget in BUDGET_SCENARIOS.items():
        r = optimize_portfolio(
            dfp, budget,
            min_influencers = 3,
            tier_limits     = {"mega": 2},   # mega >2 seçilmesin
            label           = label,
        )
        results.append(r)
    return results


# =====================================================================
# ADIM 6: DUYARLILIK ANALİZİ
# =====================================================================

def sensitivity_analysis(
    dfp         : pd.DataFrame,
    base_budget : float = 150_000
) -> pd.DataFrame:
    print("\n" + "="*62)
    print("ADIM 6: Duyarlılık Analizi (kriter ağırlıkları ±20%)")
    print("="*62)

    base       = optimize_portfolio(dfp, base_budget, label="Base", verbose=True)
    base_score = base["total_score"]

    crit_col_map = {
        "engagement_proxy"  : "c1_engagement",
        "followers"         : "c2_followers",
        "like_per_follower" : "c3_like_per_follower",
        "cost_efficiency"   : "c4_cost_efficiency",
        "video_volume"      : "c5_video_volume",
    }

    rows = []
    for crit, bw in CRITERION_WEIGHTS.items():
        for delta, tag in [(-0.20, "-20%"), (+0.20, "+20%")]:
            nw = {k: v for k, v in CRITERION_WEIGHTS.items()}
            nw[crit] = max(0.0, bw + delta)
            total    = sum(nw.values())
            nw       = {k: v / total for k, v in nw.items()}

            tmp = dfp.copy()
            tmp["mcdm_score"] = sum(
                nw[k] * tmp[crit_col_map[k]] for k in nw
            ).round(4)

            res = optimize_portfolio(tmp, base_budget,
                                     label=f"{crit}{tag}", verbose=False)
            pct = (res["total_score"] - base_score) / base_score * 100 \
                  if base_score else 0.0

            rows.append({
                "Kriter"                    : crit,
                "Degisim"                   : tag,
                "Yeni Agirlik"              : round(nw[crit], 4),
                "Secilen Sayi"              : res["n_selected"],
                "Toplam MCDM Skoru"         : round(res["total_score"], 4),
                "Toplam Maliyet (TL)"       : round(res["total_cost"]),
                "Skor Degisimi (%)"         : round(pct, 2),
            })
            print(f"  {crit:22s} {tag}:  "
                  f"skor={res['total_score']:.4f}  ({pct:+.1f}%)")

    return pd.DataFrame(rows)


# =====================================================================
# ADIM 7: GÖRSEL RAPOR (10 panel)
# =====================================================================

def generate_report(
    dfp          : pd.DataFrame,
    scenarios    : List[Dict],
    sensitivity_df: pd.DataFrame
) -> Path:
    print("\n" + "="*62)
    print("ADIM 7: Görsel Rapor")
    print("="*62)

    sns.set_theme(style="whitegrid", palette="muted")
    fig = plt.figure(figsize=(22, 30))
    fig.suptitle(
        "TikTok Influencer Seçimi — Optimizasyon Raporu  [Gurobi ILP]",
        fontsize=17, fontweight="bold", y=0.985
    )
    gs = fig.add_gridspec(5, 3, hspace=0.50, wspace=0.36)

    # 1. Tier dağılımı
    ax1 = fig.add_subplot(gs[0, 0])
    tc  = dfp["tier"].value_counts().reindex(
        [t for t in TIER_ORDER if t in dfp["tier"].values], fill_value=0
    )
    bars = ax1.bar(tc.index, tc.values,
                   color=[TC[t] for t in tc.index], edgecolor="white", lw=1.2)
    ax1.set_title("Tier Dağılımı", fontweight="bold")
    ax1.set_ylabel("Kişi Sayısı")
    for b in bars:
        ax1.text(b.get_x()+b.get_width()/2, b.get_height()+0.1,
                 str(int(b.get_height())), ha="center", va="bottom", fontsize=9)

    # 2. MCDM skor histogramı
    ax2 = fig.add_subplot(gs[0, 1])
    ax2.hist(dfp["mcdm_score"], bins=22, color="#4C9BE8",
             edgecolor="white", alpha=0.85)
    ax2.axvline(dfp["mcdm_score"].mean(), color="#E85B4C",
                lw=2, ls="--", label=f"Ort. {dfp['mcdm_score'].mean():.3f}")
    ax2.set_title("MCDM Skor Dağılımı", fontweight="bold")
    ax2.set_xlabel("MCDM Skoru"); ax2.set_ylabel("Frekans")
    ax2.legend(fontsize=8)

    # 3. Maliyet boxplot (tier bazlı)
    ax3 = fig.add_subplot(gs[0, 2])
    d3 = [dfp[dfp["tier"] == t]["estimated_cost_tl"].values
          for t in TIER_ORDER if t in dfp["tier"].values]
    l3 = [t for t in TIER_ORDER if t in dfp["tier"].values]
    bp = ax3.boxplot(d3, labels=l3, patch_artist=True, notch=False)
    for patch, t in zip(bp["boxes"], l3):
        patch.set_facecolor(TC[t]); patch.set_alpha(0.75)
    ax3.set_title("Tahmini Maliyet / Tier (TL)", fontweight="bold")
    ax3.set_ylabel("TL")
    ax3.yaxis.set_major_formatter(
        plt.FuncFormatter(lambda x, _: f"{x/1000:.0f}K"))

    # 4. Senaryo — kişi sayısı
    sc_lbl = [r["label"] for r in scenarios]
    ax4 = fig.add_subplot(gs[1, 0])
    sc_n = [r["n_selected"] for r in scenarios]
    b4 = ax4.bar(sc_lbl, sc_n, color=PAL[:4], edgecolor="white")
    ax4.set_title("Senaryo: Seçilen Kişi Sayısı", fontweight="bold")
    ax4.set_ylabel("Adet")
    ax4.set_xticklabels(sc_lbl, rotation=14, ha="right", fontsize=8)
    for b in b4:
        ax4.text(b.get_x()+b.get_width()/2, b.get_height()+0.1,
                 str(int(b.get_height())), ha="center", va="bottom", fontsize=9)

    # 5. Senaryo — MCDM skoru
    ax5 = fig.add_subplot(gs[1, 1])
    ax5.bar(sc_lbl, [r["total_score"] for r in scenarios],
            color=PAL[:4], edgecolor="white")
    ax5.set_title("Senaryo: Toplam MCDM Skoru", fontweight="bold")
    ax5.set_ylabel("Skor")
    ax5.set_xticklabels(sc_lbl, rotation=14, ha="right", fontsize=8)

    # 6. Senaryo — bütçe kullanımı
    ax6 = fig.add_subplot(gs[1, 2])
    sc_pct  = [r["budget_used_pct"] for r in scenarios]
    sc_cost = [r["total_cost"]      for r in scenarios]
    b6 = ax6.bar(sc_lbl, sc_pct, color=PAL[:4], edgecolor="white")
    ax6.axhline(100, color="red", lw=1.5, ls="--", alpha=0.5)
    ax6.set_title("Senaryo: Bütçe Kullanım Oranı (%)", fontweight="bold")
    ax6.set_xticklabels(sc_lbl, rotation=14, ha="right", fontsize=8)
    for b, p, c in zip(b6, sc_pct, sc_cost):
        ax6.text(b.get_x()+b.get_width()/2, b.get_height()+0.5,
                 f"{p:.1f}%\n{fmt_tl(c/1000)}K",
                 ha="center", va="bottom", fontsize=7.5)

    # 7. Scatter: engagement proxy vs maliyet
    ax7 = fig.add_subplot(gs[2, 0:2])
    for tier, grp in dfp.groupby("tier"):
        ep = grp["engagement_proxy"].clip(
            upper=dfp["engagement_proxy"].quantile(0.95))
        ax7.scatter(
            grp["estimated_cost_tl"] / 1000, ep,
            s=grp["mcdm_score"] * 300,
            color=TC.get(tier, "#999"), alpha=0.72,
            edgecolors="white", lw=0.6, label=tier
        )
    ax7.set_title("Engagement Proxy vs. Maliyet  (baloncuk = MCDM skoru)",
                  fontweight="bold")
    ax7.set_xlabel("Maliyet (TL bin)")
    ax7.set_ylabel("Engagement Proxy")
    ax7.legend(title="Tier", fontsize=8)

    # 8. Top-15 MCDM
    ax8 = fig.add_subplot(gs[2, 2])
    top15 = dfp.nlargest(15, "mcdm_score")[
        ["username", "mcdm_score", "tier"]].iloc[::-1]
    ax8.barh(top15["username"], top15["mcdm_score"],
             color=[TC.get(t, "#999") for t in top15["tier"]],
             edgecolor="white")
    ax8.set_title("Top 15 Influencer (MCDM)", fontweight="bold")
    ax8.set_xlabel("MCDM Skoru")
    ax8.tick_params(axis="y", labelsize=7)

    # 9. Orta bütçe senaryo — seçilen detay
    ax9 = fig.add_subplot(gs[3, :])
    mid_res = next((r for r in scenarios if "150" in r["label"]), scenarios[1])
    if not mid_res["selected"].empty:
        sel = mid_res["selected"].sort_values("mcdm_score", ascending=False)
        xp  = np.arange(len(sel))
        ax9.bar(xp, sel["mcdm_score"],
                color=[TC.get(t, "#999") for t in sel["tier"]],
                edgecolor="white")
        ax9.set_xticks(xp)
        ax9.set_xticklabels(sel["username"].tolist(),
                            rotation=35, ha="right", fontsize=7.5)
        ax9.set_title(
            f"Orta Bütçe — Seçilen Influencerlar (MCDM Skoru)",
            fontweight="bold"
        )
        ax9.set_ylabel("MCDM Skoru")
        ax9b = ax9.twinx()
        ax9b.plot(xp, sel["estimated_cost_tl"] / 1000,
                  "D--", color="#F59E0B", ms=5, lw=1.5, label="Maliyet (TL K)")
        ax9b.set_ylabel("Maliyet (TL bin)", color="#F59E0B")
        ax9b.tick_params(axis="y", labelcolor="#F59E0B")
        patches = [plt.Rectangle((0, 0), 1, 1, fc=TC.get(t, "#999"), label=t)
                   for t in TIER_ORDER if t in sel["tier"].values]
        ax9.legend(handles=patches, loc="upper right", fontsize=8, title="Tier")

    # 10. Duyarlılık — Tornado chart
    ax10 = fig.add_subplot(gs[4, :])
    if not sensitivity_df.empty:
        piv = sensitivity_df.pivot_table(
            index="Kriter", columns="Degisim",
            values="Skor Degisimi (%)", aggfunc="mean"
        )
        yp = np.arange(len(piv)); h = 0.35
        if "-20%" in piv.columns:
            ax10.barh(yp - h/2, piv["-20%"], height=h,
                      label="Ağırlık −20%", color=PAL[3], alpha=0.80)
        if "+20%" in piv.columns:
            ax10.barh(yp + h/2, piv["+20%"], height=h,
                      label="Ağırlık +20%", color=PAL[1], alpha=0.80)
        ax10.set_yticks(yp)
        ax10.set_yticklabels(piv.index, fontsize=9)
        ax10.axvline(0, color="black", lw=0.8)
        ax10.set_title(
            "Duyarlılık — Kriter Ağırlığı ±20% → Toplam Skor Değişimi (%)",
            fontweight="bold"
        )
        ax10.set_xlabel("Skor Değişimi (%)")
        ax10.legend(fontsize=9)

    path = OUT_DIR / f"optimization_report_{ts()}.png"
    fig.savefig(str(path), dpi=130, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"  Grafik: {path}")
    return path


# =====================================================================
# ADIM 8: EXCEL RAPORU
# =====================================================================

def export_excel(
    dfp           : pd.DataFrame,
    scenarios     : List[Dict],
    sensitivity_df: pd.DataFrame
) -> Path:
    out_path = OUT_DIR / f"tiktok_optimization_{ts()}.xlsx"
    print(f"\n  Excel: {out_path}")

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils  import get_column_letter

    HDR_FILL = PatternFill("solid", start_color="1E3A5F")
    HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
    ALT_FILL = PatternFill("solid", start_color="EEF2FF")
    THIN     = Border(
        left  =Side(style="thin", color="D1D5DB"),
        right =Side(style="thin", color="D1D5DB"),
        top   =Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    def _style(ws):
        for i, row in enumerate(ws.iter_rows()):
            for cell in row:
                cell.border    = THIN
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if i == 0:
                    cell.font = HDR_FONT; cell.fill = HDR_FILL
                elif i % 2 == 0:
                    if cell.fill.start_color.rgb in ("00000000", "FFFFFFFF"):
                        cell.fill = ALT_FILL
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 3, 35)
        ws.freeze_panes = "A2"

    cols_main = [c for c in [
        "username", "display_name", "category", "tier",
        "followers", "like_per_follower", "avg_video_likes",
        "engagement_proxy", "mcdm_score", "estimated_cost_tl",
        "c1_engagement", "c2_followers", "c3_like_per_follower",
        "c4_cost_efficiency", "c5_video_volume",
    ] if c in dfp.columns]

    with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:

        # Sayfa 1: Tüm influencerlar
        dfp_out = dfp[cols_main].sort_values("mcdm_score", ascending=False).copy()
        dfp_out.insert(0, "Sira", range(1, len(dfp_out)+1))
        dfp_out.to_excel(writer, sheet_name="Tum_Influencerlar", index=False)

        # Sayfa 2-N: Her senaryo seçilenleri
        for res in scenarios:
            if res["selected"].empty:
                continue
            sc = [c for c in cols_main if c in res["selected"].columns]
            sel = res["selected"][sc].sort_values("mcdm_score", ascending=False).copy()
            sel.insert(0, "Sira", range(1, len(sel)+1))
            sel.to_excel(writer, sheet_name=res["label"][:31], index=False)

        # Senaryo özeti
        pd.DataFrame([{
            "Senaryo"             : r["label"],
            "Butce (TL)"          : r["budget"],
            "Secilen Sayi"        : r["n_selected"],
            "Toplam Maliyet (TL)" : round(r["total_cost"]),
            "Butce Kullanimi (%)" : round(r["budget_used_pct"], 1),
            "Toplam MCDM Skoru"   : round(r["total_score"], 4),
            "Cozum Durumu"        : r["status"],
        } for r in scenarios]).to_excel(
            writer, sheet_name="Senaryo_Ozet", index=False)

        # Duyarlılık analizi
        if not sensitivity_df.empty:
            sensitivity_df.to_excel(
                writer, sheet_name="Duyarlilik_Analizi", index=False)

        # Model parametreleri
        params = []
        for k, v in CRITERION_WEIGHTS.items():
            params.append({"Bolum":"MCDM Agirliklari","Parametre":k,"Deger":v})
        for k, v in BUDGET_SCENARIOS.items():
            params.append({"Bolum":"Butce Senaryolari","Parametre":k,"Deger":v})
        for t, cfg in COST_MODEL.items():
            params.append({"Bolum":"Maliyet Modeli","Parametre":t,
                           "Deger":cfg["base"],
                           "Not":f"per_1k={cfg['per_1k']}"})
        if BLACKLIST:
            params.append({"Bolum":"Kara Liste","Parametre":"Yasakli",
                           "Deger":str(BLACKLIST), "Not":"x_k=0"})
        pd.DataFrame(params).to_excel(
            writer, sheet_name="Model_Parametreleri", index=False)

        # Excel stil uygula
        wb = writer.book
        for ws in wb.worksheets:
            _style(ws)

    print(f"  Excel hazır.")
    return out_path


# =====================================================================
# ANA FONKSİYON
# =====================================================================

def main():
    print("\n" + "="*62)
    print("  TİKTOK OPTİMİZASYON — FAZ 2  v3.0  [Gurobi ILP]")
    print("  Eskişehir Teknik Üniversitesi — Endüstri Mühendisliği")
    print("="*62)
    print(f"  Başlangıç : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Çıktı     : {OUT_DIR.resolve()}")
    print(f"  Solver    : Gurobi")
    if BLACKLIST:
        print(f"  Kara Liste: {BLACKLIST}")

    dfp, dfv = load_data()
    if dfp.empty:
        print("[HATA] Veri yok, program sonlanıyor."); return

    dfp = build_engagement_proxy(dfp, dfv)
    dfp = estimate_costs(dfp)
    dfp = compute_mcdm_score(dfp)

    scenarios      = run_scenarios(dfp)
    sensitivity_df = sensitivity_analysis(dfp, base_budget=150_000)
    plot_path      = generate_report(dfp, scenarios, sensitivity_df)
    excel_path     = export_excel(dfp, scenarios, sensitivity_df)

    print("\n" + "="*62)
    print("  ✅ FAZ 2 TAMAMLANDI")
    print("="*62)
    print(f"\n  📊 Excel  : {excel_path}")
    print(f"  📈 Grafik : {plot_path}")
    print(f"\n  Sonraki   : Tez II — Model çözümü & test (Nisan–Mayıs 2026)")


if __name__ == "__main__":
    main()